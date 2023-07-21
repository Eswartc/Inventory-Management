import sys
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import numbers
import pandas as pd
import datetime
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QTableWidget, QTableWidgetItem

account = []
wb = openpyxl.load_workbook("Entries.xlsx")
journal = wb["journal"]

class Ui_MainWindow(object):
        def setupUi(self, MainWindow):
                MainWindow.setObjectName("MainWindow")
                MainWindow.resize(1600, 600)
                font = QtGui.QFont()
                font.setFamily("HP Simplified Hans")
                MainWindow.setFont(font)
                MainWindow.setAutoFillBackground(False)
                self.centralwidget = QtWidgets.QWidget(MainWindow)
                self.centralwidget.setObjectName("centralwidget")
                self.textEdit = QtWidgets.QTextEdit(self.centralwidget)
                self.textEdit.setGeometry(QtCore.QRect(10, 10, 850, 500))
                font = QtGui.QFont()
                font.setFamily("HP Simplified Hans")
                font.setPointSize(10)
                self.textEdit.setFont(font)
                self.textEdit.setStyleSheet("""QTextEdit {color: #000000;background: qlineargradient(x1:1, y1:0, x2:0, y2:1, stop:0 #FFFFFF, stop:1 #787878);}""")
                self.textEdit.setAcceptRichText(True)            
                self.textEdit.setObjectName("textEdit")

                self.tabWidget = QtWidgets.QTabWidget(self.centralwidget)
                self.tabWidget.setGeometry(QtCore.QRect(880, 10, 707, 504))
                font = QtGui.QFont()
                font.setFamily("HP Simplified Hans")
                font.setBold(False)
                font.setItalic(False)
                font.setUnderline(False)
                font.setWeight(70)
                font.setStrikeOut(False)
                self.tabWidget.setFont(font)
                self.tabWidget.setMouseTracking(False)
                self.tabWidget.setStyleSheet("background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #0D47A1, stop:1 #FF5722);")

                self.tabWidget.setObjectName("tabWidget")
                self.tab = QtWidgets.QWidget()
                self.tab.setObjectName("tab")

                self.pushButton_6 = QtWidgets.QPushButton(self.centralwidget)
                self.pushButton_6.setGeometry(QtCore.QRect(325, 449, 230, 50))
                font = QtGui.QFont()
                font.setFamily("HP Simplified Hans")
                font.setPointSize(10)
                font.setBold(False)
                font.setWeight(50)
                self.pushButton_6.setFont(font)
                self.pushButton_6.setStyleSheet("QPushButton {background: #B90E0A; color: #ffffff; border: none; color: white;border-radius: 10px;}"
                "QPushButton:hover {background: #E3242B;color: white;}")


                self.pushButton_6.setObjectName("pushButton_6")

                self.pushButton = QtWidgets.QPushButton(self.tab)
                self.pushButton.setGeometry(QtCore.QRect(270, 215, 200, 50))
                font = QtGui.QFont()
                font.setFamily("HP Simplified Hans")
                font.setPointSize(10)
                font.setBold(False)
                font.setWeight(50)
                self.pushButton.setFont(font)
                self.pushButton.setStyleSheet("QPushButton {background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #0D47A1, stop:0.5 #673AB7, stop:1 #9575CD); color: white; border: none; border-radius: 10px; padding: 10px;}"
                "QPushButton:hover {background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #9575CD, stop:0.5 #673AB7, stop:1 #0D47A1);}")
                self.pushButton.setObjectName("pushButton")


                self.lineEdit_2 = QtWidgets.QLineEdit(self.tab)
                self.lineEdit_2.setGeometry(QtCore.QRect(130, 160, 115, 40))
                font = QtGui.QFont()
                font.setFamily("HP Simplified Hans")
                font.setPointSize(7)
                self.lineEdit_2.setFont(font)
                self.lineEdit_2.setStyleSheet("QLineEdit {background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #4E96B1, stop:1 #A0C5D6); border: none; border-radius: 5px; padding: 5px; color: #000000;} QLineEdit:focus {border: 1px solid #217CA3;} QLineEdit:hover {background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #6BA4C1, stop:1 #C1E4F0);}")
                self.lineEdit_2.setObjectName("lineEdit_2")


                self.lineEdit_3 = QtWidgets.QLineEdit(self.tab)
                self.lineEdit_3.setGeometry(QtCore.QRect(9, 160, 108, 40))
                font = QtGui.QFont()
                font.setFamily("HP Simplified Hans")
                font.setPointSize(7)
                self.lineEdit_3.setFont(font)
                self.lineEdit_3.setStyleSheet("QLineEdit {background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #4E96B1, stop:1 #A0C5D6); border: none; border-radius: 5px; padding: 5px; color: #000000;} QLineEdit:focus {border: 1px solid #217CA3;} QLineEdit:hover {background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #6BA4C1, stop:1 #C1E4F0);}")
                self.lineEdit_3.setObjectName("lineEdit_3")


                self.lineEdit_4 = QtWidgets.QLineEdit(self.tab)
                self.lineEdit_4.setGeometry(QtCore.QRect(259, 160, 102, 40))
                font = QtGui.QFont()
                font.setFamily("HP Simplified Hans")
                font.setPointSize(7)
                self.lineEdit_4.setFont(font)
                self.lineEdit_4.setStyleSheet("QLineEdit {background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #A0C5D6, stop:1 #F18C8E); border: 1px solid #217CA3; border-radius: 5px; padding: 5px; color: #000000;} QLineEdit:hover {background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #6BA4C1, stop:1 #FFA0A3);}")
                self.lineEdit_4.setObjectName("lineEdit_4")


                self.lineEdit_5 = QtWidgets.QLineEdit(self.tab)
                self.lineEdit_5.setGeometry(QtCore.QRect(375, 160, 198, 40))
                font = QtGui.QFont()
                font.setFamily("HP Simplified Hans")
                font.setPointSize(7)
                self.lineEdit_5.setFont(font)
                self.lineEdit_5.setStyleSheet("QLineEdit {background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #A0C5D6, stop:1 #F18C8E); border: 1px solid #217CA3; border-radius: 5px; padding: 5px; color: #000000;} QLineEdit:hover {background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #6BA4C1, stop:1 #FFA0A3);}")
                self.lineEdit_5.setObjectName("lineEdit_5")

                self.lineEdit_6 = QtWidgets.QLineEdit(self.tab)
                self.lineEdit_6.setGeometry(QtCore.QRect(585, 160, 108, 40))
                font = QtGui.QFont()
                font.setFamily("HP Simplified Hans")
                font.setPointSize(7)
                self.lineEdit_6.setFont(font)
                self.lineEdit_6.setStyleSheet("QLineEdit {background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #A0C5D6, stop:1 #F18C8E); border: 1px solid #217CA3; border-radius: 5px; padding: 5px; color: #000000;} QLineEdit:hover {background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #6BA4C1, stop:1 #FFA0A3);}")
                self.lineEdit_6.setObjectName("lineEdit_6")


                self.tabWidget.addTab(self.tab, "")
                self.tab_5 = QtWidgets.QWidget()
                self.tab_5.setObjectName("tab_5")


                self.pushButton_3 = QtWidgets.QPushButton(self.tab_5)
                self.pushButton_3.setGeometry(QtCore.QRect(270, 215, 200, 50))
                font = QtGui.QFont()
                font.setFamily("HP Simplified Hans")
                font.setPointSize(10)
                self.pushButton_3.setFont(font)
                self.pushButton_3.setStyleSheet("QPushButton {background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #0D47A1, stop:0.5 #673AB7, stop:1 #9575CD); color: white; border: none; border-radius: 10px; padding: 10px;}"
                "QPushButton:hover {background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #9575CD, stop:0.5 #673AB7, stop:1 #0D47A1);}")
                self.pushButton_3.setObjectName("pushButton_3")


                self.lineEdit = QtWidgets.QLineEdit(self.tab_5)
                self.lineEdit.setGeometry(QtCore.QRect(270, 160, 200, 40))
                font = QtGui.QFont()
                font.setFamily("HP Simplified Hans")
                font.setPointSize(7)
                self.lineEdit.setFont(font)
                self.lineEdit.setStyleSheet("QLineEdit {background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #A0C5D6, stop:1 #F18C8E); border: 1px solid #217CA3; border-radius: 5px; padding: 5px; color: #000000;} QLineEdit:hover {background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #6BA4C1, stop:1 #FFA0A3);}")
                self.lineEdit.setObjectName("lineEdit")


                self.tabWidget.addTab(self.tab_5, "")
                self.tab_2 = QtWidgets.QWidget()
                self.tab_2.setObjectName("tab_2")


                self.pushButton_2 = QtWidgets.QPushButton(self.tab_2)
                self.pushButton_2.setGeometry(QtCore.QRect(270, 215, 200, 50))
                font = QtGui.QFont()
                font.setFamily("HP Simplified Hans")
                font.setPointSize(10)
                font.setBold(False)
                font.setWeight(50)
                self.pushButton_2.setFont(font)
                self.pushButton_2.setStyleSheet("QPushButton {background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #0D47A1, stop:0.5 #673AB7, stop:1 #9575CD); color: white; border: none; border-radius: 10px; padding: 10px;}"
                "QPushButton:hover {background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #9575CD, stop:0.5 #673AB7, stop:1 #0D47A1);}")
                self.pushButton_2.setObjectName("pushButton_2")


                self.tabWidget.addTab(self.tab_2, "")
                self.tab_3 = QtWidgets.QWidget()
                self.tab_3.setObjectName("tab_3")


                self.pushButton_4 = QtWidgets.QPushButton(self.tab_3)
                self.pushButton_4.setGeometry(QtCore.QRect(270, 215, 200, 50))
                font = QtGui.QFont()
                font.setFamily("HP Simplified Hans")
                font.setPointSize(10)
                font.setBold(False)
                font.setWeight(50)
                self.pushButton_4.setFont(font)
                self.pushButton_4.setStyleSheet("QPushButton {background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #0D47A1, stop:0.5 #673AB7, stop:1 #9575CD); color: white; border: none; border-radius: 10px; padding: 10px;}"
                "QPushButton:hover {background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #9575CD, stop:0.5 #673AB7, stop:1 #0D47A1);}")
                self.pushButton_4.setObjectName("pushButton_4")


                self.tabWidget.addTab(self.tab_3, "")
                self.tab_4 = QtWidgets.QWidget()
                self.tab_4.setObjectName("tab_4")


                self.pushButton_5 = QtWidgets.QPushButton(self.tab_4)
                self.pushButton_5.setGeometry(QtCore.QRect(270, 215, 200, 50))
                font = QtGui.QFont()
                font.setFamily("HP Simplified")
                font.setPointSize(10)
                self.pushButton_5.setFont(font)
                self.pushButton_5.setStyleSheet("QPushButton {background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #0D47A1, stop:0.5 #673AB7, stop:1 #9575CD); color: white; border: none; border-radius: 10px; padding: 10px;}"
                "QPushButton:hover {background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #9575CD, stop:0.5 #673AB7, stop:1 #0D47A1);}")
                self.pushButton_5.setObjectName("pushButton_5")


                self.tabWidget.addTab(self.tab_4, "")


                self.label = QtWidgets.QLabel(self.centralwidget)
                self.label.setGeometry(QtCore.QRect(750, 200, 210, 111))
                font = QtGui.QFont()
                font.setFamily("OCR A Extended")
                font.setPointSize(10)
                font.setBold(False)
                font.setWeight(50)
                self.label.setFont(font)
                self.label.setTextFormat(QtCore.Qt.RichText)


                self.label.setObjectName("label")
                MainWindow.setCentralWidget(self.centralwidget)
                self.statusbar = QtWidgets.QStatusBar(MainWindow)
                self.statusbar.setObjectName("statusbar")
                MainWindow.setStatusBar(self.statusbar)
                self.actionView = QtWidgets.QAction(MainWindow)
                self.actionView.setObjectName("actionView")


                self.retranslateUi(MainWindow)
                self.tabWidget.setCurrentIndex(0)
                QtCore.QMetaObject.connectSlotsByName(MainWindow)
                MainWindow.setTabOrder(self.pushButton, self.pushButton_3)
                MainWindow.setTabOrder(self.pushButton_3, self.pushButton_2)
                MainWindow.setTabOrder(self.pushButton_2, self.textEdit)


                self.pushButton.clicked.connect(self.add_entry)
                self.pushButton_2.clicked.connect(self.display_journal)
                self.pushButton_3.clicked.connect(self.search)
                self.pushButton_4.clicked.connect(self.display_ledger)
                self.pushButton_5.clicked.connect(self.opening_closing)
                self.pushButton_6.clicked.connect(self.generate_journal_ledger)


        def retranslateUi(self, MainWindow):
                _translate = QtCore.QCoreApplication.translate
                MainWindow.setWindowTitle(_translate("MainWindow", "MainWindow"))
                self.textEdit.setPlaceholderText(_translate("MainWindow", "You can view all the contents here!"))
                self.pushButton.setText(_translate("MainWindow", "Add Entry"))
                self.lineEdit_2.setPlaceholderText(_translate("MainWindow", "Account Number"))
                self.lineEdit_3.setPlaceholderText(_translate("MainWindow", "Name"))
                self.lineEdit_4.setPlaceholderText(_translate("MainWindow", "DD-MM-YYYY"))
                self.lineEdit_5.setPlaceholderText(_translate("MainWindow", "Description"))
                self.lineEdit_6.setPlaceholderText(_translate("MainWindow", "Cost"))
                self.tabWidget.setTabText(self.tabWidget.indexOf(self.tab), _translate("MainWindow", "Add Entry"))
                self.pushButton_3.setText(_translate("MainWindow", "Search"))
                self.lineEdit.setPlaceholderText(_translate("MainWindow", "Account Number"))
                self.tabWidget.setTabText(self.tabWidget.indexOf(self.tab_5), _translate("MainWindow", "Search"))
                self.pushButton_2.setText(_translate("MainWindow", "Display Journal"))
                self.tabWidget.setTabText(self.tabWidget.indexOf(self.tab_2), _translate("MainWindow", "Display Journal"))
                self.pushButton_4.setText(_translate("MainWindow", "Display Ledger"))
                self.tabWidget.setTabText(self.tabWidget.indexOf(self.tab_3), _translate("MainWindow", "Display Ledger"))
                self.pushButton_5.setText(_translate("MainWindow", "Opening & Closing"))
                self.tabWidget.setTabText(self.tabWidget.indexOf(self.tab_4), _translate("MainWindow", "Opening and Closing"))
                self.pushButton_6.setText(_translate("MainWindow", "Generate Journal & Ledger"))
                self.actionView.setText(_translate("MainWindow", "View"))
                

        def validate_accno_name(self, accno, name):
            flag = 0
            for i in range(1, 51):
                if journal.cell(row = i, column = 1).value == accno:
                    flag = 1
                    x = journal.cell(row = i, column = 2).value
                    break
            if flag == 1:
                return name == x
            return True

        def add_entry(self):
            journal = wb["journal"]
            try:
                name = self.lineEdit_3.text()
                accno = int(self.lineEdit_2.text())
                date = self.lineEdit_4.text()
                description = self.lineEdit_5.text()
                debit = int(self.lineEdit_6.text())
                for e in range(1, 51):
                    if journal.cell(row = e, column = 1).value == None:
                        break
                if not self.validate_accno_name(accno, name):
                    msg = f'<span style = "color:red;">exists</span>'
                    self.textEdit.setText("Account number "+str(accno)+" already "+msg+", and UserName isn't matching!")
                    self.lineEdit_2.clear()
                    self.lineEdit_3.clear()
                    self.lineEdit_4.clear()
                    self.lineEdit_5.clear()
                    self.lineEdit_6.clear()
                else:
                    journal.cell(row = e, column = 1).value = accno
                    journal.cell(row = e, column = 2).value = name
                    journal.cell(row = e, column = 3).value = date
                    journal.cell(row = e, column = 4).value = description
                    journal.cell(row = e, column = 5).value = debit
                    self.textEdit.setText("Transactions added")
                    wb.save("Entries.xlsx")
                    with open("transactions.txt", "a") as fileob:
                        fileob.write(str(accno)+","+name+","+date+","+description+","+str(debit)+"\n")

                    self.lineEdit_2.clear()
                    self.lineEdit_3.clear()
                    self.lineEdit_4.clear()
                    self.lineEdit_5.clear()
                    self.lineEdit_6.clear()
                    self.make_ledger()
                
            except:
                self.lineEdit_2.clear()
                self.lineEdit_3.clear()
                self.lineEdit_4.clear()
                self.lineEdit_5.clear()
                self.lineEdit_6.clear()
                fields = f'<span style="color: red;">fields</span>'
                self.textEdit.setText("Please fill all the "+fields+ " with approprtiate values")
                

        def search(self):
            try:
                accno = int(self.lineEdit.text())
                headers = ["AccNo", "Name", "Date", "Description", "Debit"]
                srch = pd.read_excel("Entries.xlsx", header=None, names=headers, sheet_name="journal")
                srch["Date"] = pd.to_datetime(srch["Date"], format="%d-%m-%Y").dt.date.astype(str)
                srch["Date"] = pd.to_datetime(srch["Date"]).dt.strftime("%d-%m-%Y")
                if accno not in srch["AccNo"]:
                    x = f'<span style =  "color :red;">{str(accno)}</span>'
                    text = "Transactions with account number "+x+" not found!"
                    self.textEdit.setText(text)
                    self.lineEdit.clear()
                else:
                    found = '<span style="color: blue;font-size :25px;">Found</span>'
                    self.textEdit.setText(found)
                    self.textEdit.append("")

                    table1 = ('\t\t'.join(headers[:3])) + "\t\t\t"
                    table1 += ('\t'.join(headers[3:]))
                    table1 = f'<pre><span style="color: red; font-family: HP Simplified Hans">{table1}</span></pre>'
                    self.textEdit.append(table1)  # Append table1 without adding a newline

                    table = '\n'.join('\t'.join(f'{value:<24}' for value in row) for row in srch.itertuples(index=False) if row[0] == accno)
                    self.textEdit.append(table)
                    self.lineEdit.clear()
            except:
                empty = f'<span style="color: red;">empty</span>'
                self.textEdit.setText("Search field cannot be "+empty+ "!")
                self.lineEdit.clear()

        def display_journal(self):
            if journal.cell(row = 1, column = 1).value != None:
                jd = f'<span style = "color:blue;font-size:25px;">Journal Data</span>'
                self.textEdit.setText(jd)
                self.textEdit.append("")
                headers = ["AccNo", "Name", "Date", "Description", "Debit"]
                df = pd.read_excel("Entries.xlsx", header=None, names=headers, sheet_name="journal")
                df["Date"] = pd.to_datetime(df["Date"], format="%d-%m-%Y").dt.date.astype(str)
                df["Date"] = pd.to_datetime(df["Date"]).dt.strftime("%d-%m-%Y")
                table1 = ('\t\t'.join(headers[:3])) + "\t\t\t"
                table1 += ('\t'.join(headers[3:]))
                table1 = f'<pre><span style="color: red; font-family: HP Simplified Hans">{table1}</span></pre>'
                table1+="\n"
                self.textEdit.append(table1)
                table = '\n'.join('\t'.join(f'{value:<24}' for value in row) for row in df.itertuples(index=False))
                self.textEdit.append(table)

            else:
                empty = '<span style="color: red;">empty</span>'
                self.textEdit.setText("Journal file is "+empty+"!")


        def make_ledger(self):
            ledger = wb["ledger"]
            ledger.delete_rows(1, 50)
            wb.save("Entries.xlsx")
            df = pd.read_excel("Entries.xlsx", header=None)
            df = df.sort_values(by = [0,2], ascending=True)
            try:
                for i in range(51):
                    for j in range(5):
                        val = df.iloc[i, j]
                        ledger.cell(row = i+1, column = j+1).value = (val)
            except:
                    pass
            wb.save("Entries.xlsx")


        def display_ledger(self):
            if journal.cell(row = 1, column = 1).value != None:
                ld = f'<span style = "color:blue;font-size:25px;">Ledger Data</span>'
                self.textEdit.setText(ld)
                self.textEdit.append("")
                headers = ["AccNo", "Name", "Date", "Description", "Debit"]
                df = pd.read_excel("Entries.xlsx", header=None, names=headers, sheet_name="ledger")
                df["Date"] = pd.to_datetime(df["Date"], format="%d-%m-%Y").dt.date.astype(str)
                df["Date"] = pd.to_datetime(df["Date"]).dt.strftime("%d-%m-%Y")
                table1 = ('\t\t'.join(headers[:3])) + "\t\t\t"
                table1 += ('\t'.join(headers[3:]))
                table1 = f'<pre><span style="color: red; font-family: HP Simplified Hans">{table1}</span></pre>'
                table1+="\n"
                self.textEdit.append(table1)
                table = '\n'.join('\t'.join(f'{value:<24}' for value in row) for row in df.itertuples(index=False))
                self.textEdit.append(table)

            else:
                empty = '<span style="color: red;">empty</span>'
                self.textEdit.setText("Ledger file is "+empty+"!")

        def generate_journal_ledger(self):
            
            if(journal.cell(row = 1, column = 1).value == None):
                empty = f'<span style="color: red;">empty</span>'
                self.textEdit.setText("Journal file is "+empty+", cannot generate files")
            else:
                with open("journal.txt", "w") as fileob:
                    for i in range(1, journal.max_row):
                        if(journal.cell(row = i, column = 1).value != None):
                            fileob.write(','.join(str(journal.cell(row=i, column=j).value) for j in range(1, 6)) + "\n")

                    
                with open("ledger.txt", "w") as fileob:
                    self.make_ledger()
                    ledger = wb["ledger"]
                    for i in range(1, journal.max_row):
                        if(ledger.cell(row = i, column = 1).value != None):
                            fileob.write(','.join(str(ledger.cell(row=i, column=j).value) for j in range(1, 6)) + "\n")


                msg = f'<span style = "color: red;">Journal</span>'
                msg1 = f'<span style = "color: red;">Ledger</span>'
                self.textEdit.setText(msg+" & "+msg1+" are created!");

        def get_total(self, accno):
            ledger = wb["ledger"]
            res = 0
            for e in range(1, ledger.max_row):
                if ledger.cell(row = e, column = 1).value == accno:
                    res+=ledger.cell(row = e, column = 5).value
            return res


        def display_opening(self, accno):
            ledger = wb["ledger"]
            for i in range(1, 51):
                if ledger.cell(row = i, column = 1).value == accno:
                    x = f'<span style="color: black;">{" ".join([str(ledger.cell(row=i, column=1).value), str(ledger.cell(row=i, column=2).value)])}</span>'
                    return x


        def display_content(self, accno):
            ledger = wb["ledger"]
            for i in range(1, 51):
                if ledger.cell(row = i, column = 1).value == accno:
                    self.textEdit.append("\t".join(map(str, ["",ledger.cell(row=i, column=3).value, ledger.cell(row=i, column=4).value, ledger.cell(row=i, column=5).value])))


        def opening_closing(self):
            ledger = wb["ledger"]
            if ledger.cell(row = 1, column =1).value == None:
                empty = '<span style="color: red;">empty</span>'
                self.textEdit.setText("Ledger file is "+empty+"!")
            else:
                msg = '<span style = "color:blue;font-size:25px">Opening  and  Closing  Statements</span>'
                self.textEdit.setText(msg)
                self.textEdit.append("")
                ledger = wb["ledger"]
                accounts = []
                for i in range(1, 51):
                    x = journal.cell(row = i, column = 1).value
                    if x!=None:
                        accounts.append(x)
                accounts.sort()
                accounts = set(accounts)
                for acc in accounts:
                    self.textEdit.append(self.display_opening(acc))
                    self.display_content(acc)
                    amt = f'<span style = "color:blue;">{str(self.get_total(acc))}</span>'
                    x = f'<span style="color: red;">{(" ".join(["Closing amount:", amt]))}</span>'
                    self.textEdit.append(x)
                    self.textEdit.append("\n")

if __name__ == "__main__":
        import sys
        app = QtWidgets.QApplication(sys.argv)
        MainWindow = QtWidgets.QMainWindow()
        ui = Ui_MainWindow()
        ui.setupUi(MainWindow)
        MainWindow.show()
        sys.exit(app.exec_())
