import sys
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import numbers
import pandas as pd
import datetime

wb = openpyxl.load_workbook("Book1.xlsx")
journal = wb["journal"]
accounts = []

def is_present(accno):
    return accno in accounts

def search(accno):
    print("Your account details are:")
    headers = ["AccNo", "Name", "Date", "Description", "Debit"]

    max_widths = [len(header) for header in headers]
    rows = []

    for i in range(1, journal.max_row + 1):
        if journal.cell(row=i, column=1).value == accno:
            row = [str(journal.cell(row=i, column=j).value) for j in range(1, journal.max_column + 1)]
            rows.append(row)
            for j, value in enumerate(row):
                max_widths[j] = max(max_widths[j], len(value))

    for i, header in enumerate(headers):
        print(f"{header:{max_widths[i]}}", end="\t\t")
    print()

    for row in rows:
        for i, value in enumerate(row):
            print(f"{value:{max_widths[i]}}", end="\t\t")
        print()

def add_entry(accno, name,date, description, debit):
    global accounts
    for e in range(1, 51):
        if journal.cell(row = e, column = 1).value == None:
            break
    journal.cell(row = e, column = 1).value = accno
    journal.cell(row = e, column = 2).value = name
    journal.cell(row = e, column = 3).value = date
    journal.cell(row = e, column = 4).value = description
    journal.cell(row = e, column = 5).value = debit
    accounts.append(accno)
    wb.save("Book1.xlsx")
            
def display_journal():
    print("Journal data:")
    headers = ["AccNo", "Name","Date", "Description", "Debit"]
    df = pd.read_excel("Book1.xlsx", header=None, names=headers)
    df["Date"] = pd.to_datetime(df["Date"], format="%d-%m-%Y").dt.date.astype(str)
    df["Date"] = pd.to_datetime(df["Date"]).dt.strftime("%d-%m-%Y")
    

    max_widths = [max(df[col].astype(str).str.len().max(), len(col)) for col in df.columns]    
    for i, header in enumerate(headers):
        print(f"{header:{max_widths[i]}}", end="\t\t")
    print()
    
    for index, row in df.iterrows():
        for i, val in enumerate(row):
            print(f"{val:{max_widths[i]}}", end="\t\t")
        print() 
    
def make_ledger():
    ledger = wb["ledger"]
    df = pd.read_excel("Book1.xlsx", header=None)
    df = df.sort_values(by = [0,2])
    try:
        for i in range(51):
            for j in range(5):
                val = df.iloc[i, j]
                ledger.cell(row = i+1, column = j+1).value = (val)
    except:
        pass
    wb.save("Book1.xlsx")

def display_ledger():
    make_ledger()
    print("Ledger data:")
    headers = ["AccNo", "Name","Date", "Description", "Debit"]
    df = pd.read_excel("Book1.xlsx", header=None, names=headers, sheet_name="ledger")
    df["Date"] = pd.to_datetime(df["Date"], format="%d-%m-%Y").dt.date.astype(str)
    df["Date"] = pd.to_datetime(df["Date"]).dt.strftime("%d-%m-%Y")

    max_widths = [max(df[col].astype(str).str.len().max(), len(col)) for col in df.columns]    
    for i, header in enumerate(headers):
        print(f"{header:{max_widths[i]}}", end="\t\t")
    print()
    
    for index, row in df.iterrows():
        for i, val in enumerate(row):
            print(f"{val:{max_widths[i]}}", end="\t\t")
        print()

def get_total(accno):
    ledger = wb["ledger"]
    res = 0
    for e in range(1, ledger.max_row):
        if ledger.cell(row = e, column = 1).value == accno:
            res+=ledger.cell(row = e, column = 5).value
    return res

def display_opening(accno):
    ledger = wb["ledger"]
    for i in range(1, 51):
        if ledger.cell(row = i, column = 1).value == accno:
            print(ledger.cell(row = i, column = 1).value, "\t", ledger.cell(row = i, column = 2).value)
            break

def display_content(accno):
    ledger = wb["ledger"]
    for i in range(1, 51):
        if ledger.cell(row = i, column = 1).value == accno:
            print("\t",ledger.cell(row = i, column = 3).value, "\t", ledger.cell(row = i, column = 4).value, "\t", ledger.cell(row = i, column = 5).value)

def opening_closing():
    ledger = wb["ledger"]
    #global accounts
    accounts = [3,2,1]
    accounts.sort()
    accounts = set(accounts)
    for acc in accounts:
        display_opening(acc)
        display_content(acc)
        print("closing balance: ", get_total(acc))
        print()

